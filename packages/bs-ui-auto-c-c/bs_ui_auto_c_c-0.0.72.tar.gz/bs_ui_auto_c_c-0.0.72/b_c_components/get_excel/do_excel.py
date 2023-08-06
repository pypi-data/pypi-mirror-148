
from openpyxl import load_workbook
from collections import namedtuple


class do_excel:
    """
    excel读取
    """
    cases_list = list()

    def __init__(self, r_filename, r_sheet_name=''):
        try:
            self.r_filename = r_filename
            self.r_sheet_name = r_sheet_name
            rb = load_workbook(self.r_filename)
            if r_sheet_name == '':
                self.ws = rb.active
            else:
                self.ws = rb[self.r_sheet_name]

        except Exception as e:
            pass

    def new_namedtuple(self, namedtuple_name, col, namedtuple_min_row=None,
                       namedtuple_max_row=None, namedtuple_min_col=None, namedtuple_max_col=None):
        return self.new_namedtuple_class(namedtuple_name, col, self.ws, namedtuple_min_row,
                                         namedtuple_max_row, namedtuple_min_col, namedtuple_max_col)

    def read_excel(
            self,
            is_namedtuple=1,
            min_row=None,
            max_row=None,
            min_col=None,
            max_col=None,
            **kwargs):

        self.object_excel = self.ws.iter_rows(
            min_row=min_row,
            max_row=max_row,
            min_col=min_col,
            max_col=max_col,
            values_only=True)
        object_excel_list = list(self.object_excel) if is_namedtuple == 0 else self.aa(**kwargs)
        return object_excel_list

    def aa(self, namedtuple_name, col, namedtuple_min_row, namedtuple_max_row, namedtuple_min_col, namedtuple_max_col):
        Cases = self.new_namedtuple(namedtuple_name, col,
                                    namedtuple_min_row, namedtuple_max_row, namedtuple_min_col, namedtuple_max_col).Cases
        for data in self.object_excel:
            do_excel.cases_list.append(Cases(*data))
        return do_excel.cases_list

    class new_namedtuple_class:

        def __init__(
                self,
                namedtuple_name,
                col,
                ws,
                min_row,
                max_row,
                min_col,
                max_col):
            self.col = col
            sheet_head_tuple = tuple(ws.iter_rows(min_row=min_row, max_row=max_row,
                                                  min_col=min_col, max_col=max_col, values_only=True))[col]
            self.Cases = namedtuple(namedtuple_name, sheet_head_tuple)


if __name__ == '__main__':
    # 例：
    file = '/Users/sijunji/ProjectFile/PYObject/TMT_QA_UI_Test/assess_new_m/static/data/casedata.xlsx'
    a = do_excel(
        r_filename=file,
        r_sheet_name='prod'
    )
    b = a.read_excel(
        is_namedtuple=1,
        min_row=2,
        max_col=2,
        namedtuple_name='Cases',
        col=0,
        namedtuple_min_row=0,
        namedtuple_max_row=1,
        namedtuple_min_col=0,
        namedtuple_max_col=2)
    print(b)
# 生成一个do_excel实例，传入文件地址和sheet名称（可为空，为空取第一个），文件中只能传入xlsx格式的文件，xls暂不支持，手动转一下格式
# read_ excel
# 使用do_excel实例调用，当is_namedtuple=1时，需要传入命名元组的相关信息，不为1时，只需传入excel的行列的开始截止位置
