import openpyxl
from py3db.mysql import MySql
from os.path import join
import json
from datetime import datetime, timedelta
from .db import Db


class Scheduler(Db):

    def __init__(self, choose_db: str = "CN", group="A"):
        self.choose_db = choose_db.upper()
        self.pub_code = None
        self.job_group = group
        self.url_type = "0"
        self.spider_id = "customlink"
        self.url_format = "0"
        self.feed_enable = "1"
        self.content_topic = None
        self.cutoff_pub_date = None
        self.job_id = None
        self.attrs = {"type": "listing"}
        self.crawling_config = {"reqConfigs": []}
        self.output_topic = "DC-all-parser.comment.crawler.output.new"
        self.proxy_key = {"CN": "app", "US": "hk"}.get(self.choose_db)
        self.SCHED_NAME = "spiderScheduler"
        self.TIME_ZONE_ID = "Asia/Shanghai"
        self.PRIORITY = 5
        self.TRIGGER_STATE = "WAITING"
        self.TRIGGER_TYPE = "CRON"
        self.END_TIME = "0"
        self.MISFIRE_INSTR = "0"
        self.PREV_FIRE_TIME = -1
        self.JOB_CLASS_NAME = "com.wisers.spider.quartz.jobs.DbCrawlerJob"
        self.JOB_DATA = str()
        self.IS_DURABLE = 0
        self.IS_NONCONCURRENT = 1
        self.IS_UPDATE_DATA = 1
        self.REQUESTS_RECOVERY = 0
        self.read_db_config(choose_db=self.choose_db)
        self.encoding_item = dict()
        self.conn = self.connect()

    def set_group(self, group='A'):
        self.job_group = group

    def connect(self):
        return MySql(
            self.db_config.get("host"),
            self.db_config.get("port"),
            self.db_config.get("user"),
            self.db_config.get("password"),
            self.db_config.get("db")
        )

    def output_sql(self):
        self.conn.output_sql()

    def check_pub(self, pub_code):
        if self.conn.select("crawling_pub", condition_dict={'pub_code': pub_code, "proxy_key": self.proxy_key}):
            print(
                f"该配置已插入过! table:crawling_pub,pub_code:{pub_code},proxy_key={self.proxy_key}")
            return True
        else:
            return False

    def check_feed(self, pub_code, url, section, crawling_config):
        if self.conn.select('crawling_feeds', condition_dict={
            "pub_code": pub_code,
            "url": url,
            "url_type": self.url_type,
            "section": section,
            "crawling_config": crawling_config,
            "attrs": self.attrs
        }):
            print(
                f"该配置已插入过! table:crawling_feeds,pub_code:{pub_code},url={url},section={section}")
            return True
        else:
            return False

    def check_job_detail(self, pub_code):
        if self.conn.select("qrtz_job_details", condition_dict={
            "JOB_NAME": pub_code,
            "JOB_GROUP": self.job_group
        }):
            print(
                f"该配置已插入过! table:qrtz_job_details,pub_code:{pub_code},job_group={self.job_group}")
            return True
        else:
            return False

    def check_cron(self, pub_code):
        if self.conn.select("qrtz_cron_triggers", condition_dict={
            "TRIGGER_NAME": pub_code,
            "TRIGGER_GROUP": self.job_group
        }):
            print(
                f"该配置已插入过! table:qrtz_cron_triggers,pub_code:{pub_code},group={self.job_group}")
            return True
        else:
            return False

    def check_trigger(self, pub_code):
        if self.conn.select("qrtz_triggers", condition_dict={
            "TRIGGER_NAME": pub_code,
            "TRIGGER_GROUP": self.job_group
        }):
            print(
                f"该配置已插入过! table:qrtz_triggers,pub_code:{pub_code},group={self.job_group}")
            return True
        else:
            return False

    def check_url_format(self, pub_code, url_type=0):
        if self.conn.select("crawling_url_format", condition_dict={
            "pub_code": pub_code,
            "url_type": url_type
        }):
            print(
                f"该配置已插入过! table:crawling_url_format,pub_code:{pub_code},url_type={url_type}")
            return True
        else:
            return False

    def set_encoding(self, encoding, crawling_scope):
        self.encoding_item[crawling_scope] = encoding

    def add_pub(self, pub_code):
        update_time = (datetime.now()).strftime("%Y-%m-%d %H:%M:%S")
        if not self.check_pub(pub_code):
            self.conn.insert_one("crawling_pub",
                                 [None, pub_code, pub_code,
                                  'News', 1, self.proxy_key, None,
                                  update_time, update_time, None, None])

    def add_feed_pub(self, excel_name="insertFeeds.xlsx"):
        wb = openpyxl.load_workbook(join(self.base_dir, excel_name))
        ws = wb.active
        i = 0
        for r in range(2, ws.max_row + 1):
            pub_code = ws.cell(r, 1).value
            if pub_code is None:
                continue
            elif not self.pub_code:
                self.pub_code = pub_code
            update_time = (datetime.now()).strftime("%Y-%m-%d %H:%M:%S")
            if pub_code and i == 0:
                i += 1
                self.add_pub(pub_code)

            url = ws.cell(r, 2).value
            section = ws.cell(r, 3).value
            listing_headers = ws.cell(r, 4).value
            content_headers = ws.cell(r, 5).value
            listing_post_data = ws.cell(r, 6).value
            attrs_value = ws.cell(r, 7).value
            crawling_config = self.get_crawling_config(listing_headers, content_headers, listing_post_data)
            if attrs_value:
                print(attrs_value)
                self.attrs = dict(json.loads(attrs_value))
            self.attrs = json.dumps(self.attrs, separators=(',', ':'), ensure_ascii=False)
            if self.check_feed(pub_code, url, section, crawling_config):
                continue
            if self.choose_db == "CN":
                self.conn.insert_one(
                    "crawling_feeds",
                    [
                        None, pub_code, self.job_group, url, self.url_type, self.spider_id,
                        self.url_format, self.feed_enable, section, crawling_config,
                        self.attrs, self.content_topic, self.output_topic, update_time,
                        update_time, self.cutoff_pub_date, self.job_id
                    ])
            elif self.choose_db == "US":
                self.conn.insert_one(
                    "crawling_feeds",
                    [
                        None, pub_code, self.job_group, url, self.url_type, self.spider_id,
                        self.url_format, section, crawling_config,
                        self.attrs, self.content_topic, self.output_topic, update_time,
                        update_time, self.cutoff_pub_date
                    ])
            self.attrs = {"type": "listing"}
            self.crawling_config = {"reqConfigs": []}

    def add_url_format(self, pub_code, url_type=0, format_data=None):
        update_time = (datetime.now()).strftime("%Y-%m-%d %H:%M:%S")
        if not self.check_url_format(pub_code, url_type):
            self.conn.insert_one('crawling_url_format',
                                 [None, pub_code, url_type, format_data, 'template', 'listing', update_time,
                                  update_time])

    def get_crawling_config(
            self,
            listing_headers=None,
            content_headers=None,
            listing_post_data=None
    ):
        if not listing_headers and not content_headers and not listing_post_data:
            if self.crawling_config.get("reqConfigs") or self.crawling_config.get("reqConfigs") == []:
                del self.crawling_config['reqConfigs']

        if listing_post_data:
            self.crawling_config['reqConfigs'].append({
                "key": "method",
                "value": "post",
                "crawlingScope": "listing"
            })
            self.crawling_config['reqConfigs'].append({
                "key": "body",
                "value": json.loads(listing_post_data),
                "crawlingScope": "listing"
            })
        if listing_headers:
            print(listing_headers)
            self.crawling_config['reqConfigs'].append({
                "key": "headers",
                "value": json.loads(listing_headers),
                "crawlingScope": "listing"
            })
        if content_headers:
            self.crawling_config['reqConfigs'].append({
                "key": "headers",
                "value": json.loads(content_headers),
                "crawlingScope": "content"
            })
        for key, value in self.encoding_item.items():
            if self.crawling_config.get("encodingConfigs"):
                self.crawling_config.get("encodingConfigs").append({"charset": value, "crawlingScope": key})
            else:
                self.crawling_config["encodingConfigs"] = [{"charset": value, "crawlingScope": key}]

        if len(self.crawling_config):
            return json.dumps(self.crawling_config, separators=(',', ':'), ensure_ascii=False)
        else:
            return None

    def check_cron_valid(self, days=0, hours=0, minutes=0, pub_code=None):
        is_valid = False
        if not pub_code:
            print("pubcode为空!")
        elif not isinstance(days, int) or not isinstance(minutes, int) or not isinstance(hours, int):
            print("时间必须为整数")
        elif days < 0 or hours < 0 or minutes < 0:
            print("时间不能为负数")
        else:
            is_valid = True
        return is_valid

    @classmethod
    def cron_to_higher_level(cls, days=0, hours=0, minutes=0):
        if minutes % 60 == 0 and minutes >= 60:
            hours = minutes // 60
            minutes = 0
        if hours % 24 == 0 and hours >= 24:
            days = hours // 24
            hours = 0
        return days, hours, minutes

    def set_cron(self, days=0, hours=0, minutes=0, pub_code=None):
        self.pub_code = pub_code if pub_code else self.pub_code
        days, hours, minutes = self.cron_to_higher_level(days, hours, minutes)
        if self.check_cron_valid(days=days, hours=hours, minutes=minutes, pub_code=self.pub_code):
            cron = f"0 0 0 */{days} *  ?" if days else None
            cron = f"0 0 */{hours} * * ?" if hours else cron
            cron = f"0 */{minutes} * * * ?" if minutes else cron
            seq_time = ((days * 24 + hours) * 60 + minutes) * 60
            now_time, next_time = self.get_next_time(seq_time)
            if not self.check_job_detail(pub_code):
                gmt_format = '#  #%a %b %d %H:%M:%S CST %Y  '
                job_data = datetime.now().strftime(gmt_format)
                self.conn.insert_one("qrtz_job_details", [
                    self.SCHED_NAME, pub_code, self.job_group,
                    None, self.JOB_CLASS_NAME, self.IS_DURABLE,
                    self.IS_NONCONCURRENT, self.IS_UPDATE_DATA,
                    self.REQUESTS_RECOVERY, job_data]
                                     )
            if not self.check_trigger(pub_code):
                self.conn.insert_one("qrtz_triggers",
                                     [self.SCHED_NAME, self.pub_code, self.job_group,
                                      self.pub_code, self.job_group, None, next_time,
                                      self.PREV_FIRE_TIME, self.PRIORITY, self.TRIGGER_STATE,
                                      self.TRIGGER_TYPE, now_time, self.END_TIME, None,
                                      self.MISFIRE_INSTR, self.JOB_DATA
                                      ])
            if not self.check_cron(pub_code):
                self.conn.insert_one("qrtz_cron_triggers", [
                    self.SCHED_NAME, pub_code, self.job_group, cron, self.TIME_ZONE_ID])

    @staticmethod
    def get_next_time(seq_time):
        now_time = datetime.now()
        add_second = seq_time - (now_time.hour * 3600 +
                                 now_time.minute * 60 + now_time.second) % seq_time
        next_time = now_time + timedelta(seconds=add_second)
        return int(now_time.timestamp()) * 1000, int(next_time.timestamp()) * 1000
