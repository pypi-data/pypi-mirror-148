"""
Creates Excel file in raw/xlsx having two sheets 'data' and 'metadata'
based on data of the passed language ID.
If passed language ID is 'template' empty sheets of all parameter will be generated.
"""
import pathlib
import collections

import openpyxl
import pycldf
import pylexibank
from cldfbench.cli_util import with_dataset, add_dataset_spec

from ..numerals_utils import XLSX_LABELS


def register(parser):
    add_dataset_spec(parser)
    parser.add_argument(
        '--lang-id',
        type=str,
        help='CDLF Language_ID',
        default=None,
    )


def run(args):
    with_dataset(args, create_language_sheet)


def create_language_sheet(dataset, args):
    assert args.lang_id
    is_template = False
    ds = dataset.cldf_reader()

    desired_lang_id = args.lang_id
    if desired_lang_id.lower() == 'template':
        desired_lang_id = '{glottocode}-{x}'
        is_template = True
        lg = collections.defaultdict(str)
    else:
        lg = ds.get_row('LanguageTable', desired_lang_id)
        assert lg
        args.log.info('creating Excel file for "{}" ...'. format(lg['Name']))

    xlsx_dir = dataset.raw_dir / 'xlsx'
    xlsx_dir.mkdir(exist_ok=True)
    xlsx_path = xlsx_dir / 'numerals-{}.xlsx'.format(desired_lang_id)
    if xlsx_path.exists():  # pragma: no cover
        args.log.error('{} already exists'.format(xlsx_path))
        return

    header = [XLSX_LABELS['param'], XLSX_LABELS['form'],
              XLSX_LABELS['form_comment'], XLSX_LABELS['loan'], XLSX_LABELS['other_form']]
    header_letters = [openpyxl.utils.get_column_letter(i + 1) for i in range(len(header))]

    header_font = openpyxl.styles.Font(bold=True, size=16)
    header_alignment = openpyxl.styles.Alignment(horizontal='center')
    right_alignment = openpyxl.styles.Alignment(horizontal='right', wrap_text=True, vertical='top')
    md_alignment = openpyxl.styles.Alignment(wrap_text=True, vertical='top')
    cell_font = openpyxl.styles.Font(size=16, name='Charis SIL')
    pattern_fill = openpyxl.styles.PatternFill("solid", fgColor="EEEEEE")
    format_text = openpyxl.styles.numbers.FORMAT_TEXT

    wb = openpyxl.Workbook()
    wdata = wb.active
    wdata.title = XLSX_LABELS['data']
    wdata.append(header)
    wdata.freeze_panes = "A2"
    for c in header_letters:
        wdata["{}1".format(c)].font = header_font
        wdata["{}1".format(c)].alignment = header_alignment

    row_cnt = 1
    found = False
    if is_template:
        target = ds['ParameterTable']
    else:
        target = ds['FormTable']
    for row in pylexibank.progressbar(target, desc='Extracting forms'):
        if is_template:
            p = row['Name']
            row = collections.defaultdict(str)
            row['Parameter_ID'] = p
        else:
            if row['Language_ID'] != desired_lang_id:
                if found:
                    break
                else:  # pragma: no cover
                    continue
        row_cnt += 1
        d = list(map(int, row['Parameter_ID'].split('-')))
        d.append(row['Form'])
        d.append(row['Comment'])
        d.append(1 if row['Loan'] else '')
        d.append(';'.join(row['Other_Form']) if isinstance(row['Other_Form'], list) else row['Other_Form'])
        wdata.append(d)
        for c in header_letters:
            wdata["{}{}".format(c, row_cnt)].font = cell_font
            wdata["{}{}".format(c, row_cnt)].number_format = format_text
            wdata["A{}".format(row_cnt)].fill = pattern_fill
            wdata["A{}".format(row_cnt)].alignment = right_alignment
        found = True

    wdata.column_dimensions['A'].width = 16
    wdata.column_dimensions['B'].width = 44
    wdata.column_dimensions['C'].width = 60
    wdata.column_dimensions['D'].width = 11
    wdata.column_dimensions['E'].width = 54

    wmeta = wb.create_sheet(title=XLSX_LABELS['metadata'])
    wmeta.append([XLSX_LABELS['glottocode'], lg['Glottocode']])
    wmeta.append([XLSX_LABELS['isocode'], lg['ISO639P3code']])
    wmeta.append([XLSX_LABELS['name'], lg['Name']])
    wmeta.append([XLSX_LABELS['name_zh'], ''])
    wmeta.append([XLSX_LABELS['sourcefile'], lg['SourceFile']])
    wmeta.append([XLSX_LABELS['base'], lg['Base']])
    wmeta.append([XLSX_LABELS['author_fam'], ''])
    wmeta.append([XLSX_LABELS['author_first'], ''])
    wmeta.append([XLSX_LABELS['author'], lg['Contributor']])
    wmeta.append([XLSX_LABELS['ref'], ''])
    wmeta.append([XLSX_LABELS['glotto_ref_id'], ''])
    wmeta.append([XLSX_LABELS['date'], ''])
    wmeta.append([XLSX_LABELS['countries'], ''])
    wmeta.append([XLSX_LABELS['other_location'], ''])
    wmeta.append([XLSX_LABELS['lg_comment'], lg['Comment']])

    for c in range(1, 16):
        wmeta["{}{}".format('A', c)].font = header_font
        wmeta["{}{}".format('A', c)].alignment = right_alignment
        wmeta["{}{}".format('A', c)].fill = pattern_fill
        wmeta["{}{}".format('B', c)].font = cell_font
        wmeta["{}{}".format('B', c)].number_format = format_text
        wmeta["{}{}".format('B', c)].alignment = md_alignment

    wmeta.column_dimensions['A'].width = 45
    wmeta.column_dimensions['B'].width = 175
    wmeta.row_dimensions[15].height = 1530
    # comment cell
    wmeta["B15"].font = openpyxl.styles.Font(size=12, name='Charis SIL')

    wb.save(xlsx_path)
