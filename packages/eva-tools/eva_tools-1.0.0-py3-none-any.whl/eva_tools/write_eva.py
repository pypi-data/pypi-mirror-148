
from .read_eva import ReadEvaNC
import netCDF4 as nc4
from datetime import datetime
import matplotlib.pyplot as plt
import numpy as np
import openpyxl
from openpyxl.chart import Reference, Series, ScatterChart
from openpyxl.drawing.line import LineProperties
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment


class WriteEvaNC:

    nc = None
    time_desc = 'Time, years'
    time_units = 'Years'
    lat_units = 'Degrees N'

    def __init__(self, filename, time, latitude, aod550, metadata, time_unit=None, lat_unit=None, time_desc=None):
        if time_unit is not None:
            self.time_units = time_unit
        if time_desc is not None:
            self.time_desc = time_desc
        if lat_unit is not None:
            self.lat_units = lat_unit
        self._create_nc_file(filename, time, latitude)
        self._write_meta_data(metadata)
        self._write_data(time, latitude, aod550)
        self.nc.close()

    def _create_nc_file(self, filename, time, latitude):
        nc = nc4.Dataset(filename, clobber=True, mode='w')
        nc.createDimension('time', len(time))
        nc.createDimension('lat', len(latitude))
        self.nc = nc

    def _write_meta_data(self, metadata):
        for key in metadata:
            self.nc.setncattr(key, metadata[key])

    def _write_data(self, time, latitude, aod550):
        # Write time data
        time_nc = self.nc.createVariable('time', 'f8', 'time')
        time_nc.setncattr('units', self.time_units)
        time_nc.setncattr('Long name', self.time_desc)
        time_nc[:] = time
        # Write latitude data
        lat_nc = self.nc.createVariable('lat', 'f8', 'lat')
        lat_nc.setncattr('units', self.lat_units)
        lat_nc.setncattr('Long name', 'Latitude')
        lat_nc[:] = latitude
        # Write aod550 data
        aod_nc = self.nc.createVariable('aod550', 'f8', ('time', 'lat'))
        aod_nc.setncattr('units', 'unitless')
        aod_nc.setncattr('Long name', 'stratospheric aerosol optical depth at 550 nm')
        aod550 = np.array(aod550)
        aod_nc[:, :] = aod550

class WriteData:
    """
    Writes data from a ReadEvaNC object to other formats.
    Currently supported formats are: Excel (.xlsx).
    """
    # Save raw input data
    data = None
    format = None
    # Save each data field
    aod550 = None
    gm_aod = None
    n3090aod = None
    n030aod = None
    s030aod = None
    s3090aod = None
    north_aod = None
    south_aod = None
    time = None
    latitude = None
    # Hold times if compressed to years
    time_year = None

    def __init__(self, data_in):
        """
        :param data_in: ReadData object containing aod550 data
        """
        # Read in data
        self.data = data_in
        self.format = data_in.format
        # Read in data fields
        self.time = list(self.data.time)
        self.aod550 = self.data.aod550[:, :]
        self.gm_aod = list(self.data.gm_aod_550)
        self.north_aod = list(self.data.hemi_aod_550[0])
        self.south_aod = list(self.data.hemi_aod_550[1])
        lat4aod = self.data.lat4_aod_550
        self.n3090aod = list(lat4aod[0])
        self.n030aod = list(lat4aod[1])
        self.s030aod = list(lat4aod[2])
        self.s3090aod = list(lat4aod[3])
        self.latitude = list(self.data.lat)

    def compress_aod_to_year(self):
        # Get starting year
        year = self.time[0] // 1
        # Create lists for storing data in each year to average
        #aod_avg = []
        gm_avg = []
        north_avg = []
        south_avg = []
        n3090_avg = []
        n030_avg = []
        s030_avg = []
        s3090_avg = []
        # Create output lists
        time_out = []
        gm_out = []
        north_out = []
        south_out = []
        n3090_out = []
        n030_out = []
        s030_out = []
        s3090_out = []

        # iterate through all time data
        for index, this_time in enumerate(self.time):
            if year != this_time // 1:
                # If in new year, save data from previous year
                # Save yearly averaged data
                time_out.append(year)
                gm_out.append(sum(gm_avg)/len(gm_avg))
                north_out.append(sum(north_avg) / len(north_avg))
                south_out.append(sum(south_avg) / len(south_avg))
                n030_out.append(sum(n030_avg) / len(n030_avg))
                n3090_out.append(sum(n3090_avg) / len(n3090_avg))
                s030_out.append(sum(s030_avg) / len(s030_avg))
                s3090_out.append(sum(s3090_avg) / len(s3090_avg))
                # Clear average lists
                gm_avg = []
                north_avg = []
                south_avg = []
                n3090_avg = []
                n030_avg = []
                s030_avg = []
                s3090_avg = []
                # Update year check
                year = this_time // 1
                # aod_avg.append(self.aod550[])
            # Add data to averaging lists for this year
            gm_avg.append(self.gm_aod[index])
            north_avg.append(self.north_aod[index])
            south_avg.append(self.south_aod[index])
            n030_avg.append(self.n030aod[index])
            n3090_avg.append(self.n030aod[index])
            s030_avg.append(self.s030aod[index])
            s3090_avg.append(self.s3090aod[index])
        # Save yearly averaged data for final year
        time_out.append(year)
        gm_out.append(sum(gm_avg) / len(gm_avg))
        north_out.append(sum(north_avg) / len(north_avg))
        south_out.append(sum(south_avg) / len(south_avg))
        n030_out.append(sum(n030_avg) / len(n030_avg))
        n3090_out.append(sum(n3090_avg) / len(n3090_avg))
        s030_out.append(sum(s030_avg) / len(s030_avg))
        s3090_out.append(sum(s3090_avg) / len(s3090_avg))

        # Save compressed data to global attributes
        self.time = time_out
        self.gm_aod = gm_out
        self.north_aod = north_out
        self.south_aod = south_out
        self.n030aod = n030_out
        self.n3090aod = n3090_out
        self.s030aod = s030_out
        self.s3090aod = s3090_out

    def toExcel(self, meta=None, source_meta=True, aod550=True, gm=True, hemi=True, lat4=True, charts=True, images=False, avg_years=True):
        """
        Writes the data contained in this object to an excel workbook, and saves it.
        :preconditions: self.data contains a ReadData object
        :return: None
        """
        if avg_years:
            self.compress_aod_to_year()

        # Open a new excel workbook
        wb = openpyxl.Workbook()
        # Remove the active sheet opened by default (all our methods create sheets).
        wb.remove(wb.active)
        # Write sheet metadata to a new sheet
        self._write_excel_meta_sheet(wb, meta)
        if source_meta:
            # Write source Metadata to a new sheet
            self._write_source_meta_sheet(wb)
        if aod550:
            # Write aod550 data to a new sheet
            self._write_aod550_sheet(wb)
        if gm:
            # Generate sheet for global mean AOD
            gm_sheet = self._write_global_mean_sheet(wb)
            # Generate chart if desired
            if charts:
                self._write_data_charts(gm_sheet=gm_sheet)
        if hemi:
            # Generate sheet for hemispheric mean AOD
            hemi_sheet = self._write_hemisphere_sheet(wb)
            # Generate chart if desired
            if charts:
                self._write_data_charts(hemi_sheet=hemi_sheet)
        if lat4:
            # Generate sheet for four latitude region mean AOD
            lat4_sheet = self._write_lat4_sheet(wb)
            # Generate chart if desired
            if charts:
                self._write_data_charts(lat4_sheet=lat4_sheet)
        if images:
            # Generate sheet for data images
            self._write_images_sheet(wb)
        # Save the completed workbook
        name_out = self.data.file[:-2] + "xlsx"
        wb.save(name_out)

    def _write_excel_meta_sheet(self, workbook, meta):
        """
        Writes metadata on the current writing process to a new workbook sheet.
        :param workbook: Reference to the openpyxl workbook we are writing to.
        :return: Reference to the new openpyxl sheet
        """
        wb = workbook
        # Create the metasheet
        meta_sheet = wb.create_sheet("Metadata")
        if meta is not None:
            # Fill metasheet with provided values
            for index, key in enumerate(meta.keys()):
                meta_sheet['A' + str(1 + index)] = key
                meta_sheet['A' + str(1 + index)].alignment = Alignment(wrap_text=True)
                meta_sheet['B' + str(1 + index)] = meta[key]
                meta_sheet['B' + str(1 + index)].alignment = Alignment(wrap_text=True)

        else:
            # Fill meta sheet with default info
            meta_sheet['A1'] = "Description:"
            meta_sheet['B1'] = "Stratospheric Aerosol Optical Depth data. This file is was created from a netcdf data file, converted to excel for data analyis. Conversion was done using the openpyxl library, with the 'data_converter.py' script."
            meta_sheet['A2'] = "Author(s):"
            meta_sheet['B2'] = "No author was provided for this file write. The data_converter.py script used for conversion was authored by C.J. Ogilvie, working for Dr. Matthew Toohey at the University of Saskatchewan."
            meta_sheet['A3'] = "Creation time:"
            meta_sheet['B3'] = datetime.now().strftime("%b %d %Y at %H:%M")
            meta_sheet['A4'] = "Tab Descriptions:"
            meta_sheet['A5'] = "Source Metadata"
            meta_sheet['B5'] = "Stores metadata attributes of the source .nc file, such as author, creation time, and references."
            meta_sheet['A6'] = "Zonal Mean"
            meta_sheet['B6'] = "Contains the mean SAOD values for every latitude region obtained by the data source, at each time value."
            meta_sheet['A7'] = "Global Mean"
            meta_sheet['B7'] = "Contains the global mean SAOD value at each point in time. Plot included."
            meta_sheet['A8'] = "Hemispheric Mean"
            meta_sheet['B8'] = "Contains the mean SAOD values for the North and South Hemispheres at each point in time. Plot included."
            meta_sheet['A9'] = "Four LatBand Mean"
            meta_sheet['B9'] = "Contains the mean SAOD values for four equal area latitude regions, being 30-90 degrees S, 0-30 S, 0-30 N, and 30-90 N."


            meta_sheet.column_dimensions['A'].width = 30
            meta_sheet.column_dimensions['B'].width = 150
            # Set all used locations to have wrapped text
            indices = ['A1', 'B1', 'A2', 'B2', 'A3', 'B3']
            for index in indices:
                meta_sheet[index].alignment = Alignment(wrap_text=True)
        return meta_sheet

    def _write_source_meta_sheet(self, workbook):
        """
        Writes metadata from the data source to a new workbook sheet.
        :param workbook: Reference to the openpyxl workbook we are writing to.
        :return:  Reference to the new openpyxl sheet
        """
        wb = workbook
        # Create the metasheet
        meta_sheet = wb.create_sheet("Source Metadata")
        meta_dict = self.data.metadata
        for index, key in enumerate(meta_dict.keys()):
            meta_sheet['A' + str(1 + index)] = key
            meta_sheet['A' + str(1 + index)].alignment = Alignment(wrap_text=True)
            meta_sheet['B' + str(1 + index)] = meta_dict[key]
            meta_sheet['B' + str(1 + index)].alignment = Alignment(wrap_text=True)
        meta_sheet.column_dimensions['A'].width = 30
        meta_sheet.column_dimensions['B'].width = 150
        return meta_sheet

    def _write_aod550_sheet(self, workbook):
        """
        Writes aod550 data to a new worksheet.
        :param workbook: Reference to openpyxl workbook we are writing to.
        :return:  Reference to the new openpyxl sheet
        """
        # Create the sheet
        aod550_sheet = workbook.create_sheet("Zonal Mean")
        # Extract needed data
        aod550 = self.aod550
        time = self.time
        latitude = self.latitude
        # First, write latitude data to header
        aod550_sheet.append(latitude)
        # Write in all aod550 data
        for index, element in enumerate(time):
            time_row = list(aod550[index, :])
            aod550_sheet.append(time_row[:])
        # Then, add spacing rows and columns for time and latitude
        aod550_sheet.insert_cols(1, amount=2)
        aod550_sheet.insert_rows(2, amount=1)
        aod550_sheet['A2'] = 'Time: (' + self.data.time_units + ')'
        aod550_sheet['A2'].alignment = Alignment(wrap_text=True)
        aod550_sheet['B1'] = 'Latitude: (' + self.data.lat_units + ')'
        aod550_sheet['B1'].alignment = Alignment(wrap_text=True)
        # Write time data to header column
        for index, element in enumerate(time, start=1):
            aod550_sheet.cell(2 + index, 1).value = element
        return aod550_sheet

    def _write_global_mean_sheet(self, workbook):
        """
        Writes global mean AOD 550 data to a new worksheet.
        :param workbook: Reference to the openpyxl workbook we are writing to.
        :return:  Reference to the new openpyxl sheet
        """
        wb = workbook
        time = self.time
        gmAOD_sheet = wb.create_sheet("Global Mean")
        # Add the data we need
        gm_aod = list(self.gm_aod)
        for index, element in enumerate(time):
            data = [element, gm_aod[index]]
            gmAOD_sheet.append(data)
        # Then add headers
        gmAOD_sheet.insert_rows(1)
        gmAOD_sheet['A1'] = "Time: (" + self.data.time_units + ')'
        gmAOD_sheet['A1'].alignment = Alignment(wrap_text=True)
        gmAOD_sheet['B1'] = "Global Mean SAOD"
        gmAOD_sheet['B1'].alignment = Alignment(wrap_text=True)
        return gmAOD_sheet

    def _write_hemisphere_sheet(self, workbook):
        """
        Writes hemispheric mean AOD 550 data to a new worksheet.
        :param workbook: Reference to the openpyxl workbook we are writing to.
        :return:  Reference to the new openpyxl sheet
        """
        # Setup the sheet
        wb = workbook
        hemi_sheet = wb.create_sheet("Hemispheric Mean")
        # Add the data we need
        time = self.time
        north_aod = self.north_aod
        south_aod = self.south_aod
        for index, element in enumerate(time):
            data = [element, north_aod[index], south_aod[index]]
            hemi_sheet.append(data)
        # Then add headers
        hemi_sheet.insert_rows(1)
        hemi_sheet['A1'] = "Time: (" + self.data.time_units + ')'
        hemi_sheet['A1'].alignment = Alignment(wrap_text=True)
        hemi_sheet['B1'] = "North Hemisphere Mean SAOD"
        hemi_sheet['B1'].alignment = Alignment(wrap_text=True)
        hemi_sheet['C1'] = "South Hemisphere Mean SAOD"
        hemi_sheet['C1'].alignment = Alignment(wrap_text=True)
        return hemi_sheet

    def _write_lat4_sheet(self, workbook):
        """
       Writes mean AOD 550 data for four equal area latitude regions to a new worksheet.
       :param workbook: Reference to the openpyxl workbook we are writing to.
       :return:  Reference to the new openpyxl sheet
       """
        # Setup the sheet
        wb = workbook
        lat4_sheet = wb.create_sheet("Four LatBand Mean")
        # Add the data we need
        time = self.time
        n3090aod = self.n3090aod
        n030aod = self.n030aod
        s030aod = self.s030aod
        s3090aod = self.s3090aod
        for index, element in enumerate(time):
            data = [element, n3090aod[index], n030aod[index], s030aod[index], s3090aod[index]]
            lat4_sheet.append(data)
        # Then add headers
        lat4_sheet.insert_rows(1)
        lat4_sheet['A1'] = "Time: (" + self.data.time_units + ')'
        lat4_sheet['A1'].alignment = Alignment(wrap_text=True)
        lat4_sheet['B1'] = "North 30-90deg Mean SAOD"
        lat4_sheet['B1'].alignment = Alignment(wrap_text=True)
        lat4_sheet['C1'] = "North 0-30deg Mean SAOD"
        lat4_sheet['C1'].alignment = Alignment(wrap_text=True)
        lat4_sheet['D1'] = "South 0-30deg Mean SAOD"
        lat4_sheet['D1'].alignment = Alignment(wrap_text=True)
        lat4_sheet['E1'] = "South 30-90deg Mean SAOD"
        lat4_sheet['E1'].alignment = Alignment(wrap_text=True)
        return lat4_sheet

    def _write_images_sheet(self, workbook):
        """
        Creates an openpyxl sheet to which images of matplotlib plots representing mean aod data are added to.
        :param workbook: The openpyxl workbook to add to
        :return: Reference to the new openpyxl sheet
        """
        wb = workbook
        image_sheet = wb.create_sheet("Data Plots")
        # Plot global mean AOD
        plt.figure(figsize=(20, 10))
        plt.clf()
        plt.plot(self.time, self.gm_aod, '-')
        plt.title("Global Mean SAOD at 550nm")
        plt.xlabel("Time (" + self.data.time_units + ")")
        plt.ylabel("Mean SAOD at lambda = 550nm")
        plt.tight_layout(pad=2)
        plt.gca().set_ylim(bottom=0)
        plt.savefig("images/gm_aod_tmp.png")
        # Plot hemispheric mean AOD
        plt.figure(figsize=(20, 10))
        plt.clf()
        plt.plot(self.time, self.north_aod, '-', label="North Hemisphere")
        plt.plot(self.time, self.south_aod, '-', label="South Hemisphere")
        plt.title("Hemispheric Mean SAOD Trend at 550nm")
        plt.xlabel("Time (" + self.data.time_units + ")")
        plt.ylabel("Mean SAOD at lambda = 550nm")
        plt.tight_layout(pad=2)
        plt.gca().set_ylim(bottom=0)
        plt.legend()
        plt.savefig("images/hemi_aod_tmp.png")
        # Plot four latitude mean AOD
        plt.figure(figsize=(20, 10))
        plt.clf()
        plt.plot(self.time, self.n3090aod, '-', label="30-90 deg North")
        plt.plot(self.time, self.n030aod, '-', label="0-30 deg North")
        plt.plot(self.time, self.s030aod, '-', label="0-30 deg South")
        plt.plot(self.time, self.s3090aod, '-', label="30-90 deg South")
        plt.title("Four LatBand SAOD Trend at 550nm")
        plt.xlabel("Time (" + self.data.time_units + ")")
        plt.ylabel("Mean SAOD at lambda = 550nm")
        plt.tight_layout(pad=2)
        plt.gca().set_ylim(bottom=0)
        plt.legend()
        plt.savefig("images/lat4_aod_tmp.png")
        # Add images to the sheet
        image_sheet.add_image(Image('images/gm_aod_tmp.png'), 'A1')
        image_sheet.add_image(Image('images/hemi_aod_tmp.png'), 'E1')
        image_sheet.add_image(Image('images/lat4_aod_tmp.png'), 'I1')
        return image_sheet

    def _write_data_charts(self, gm_sheet=None, hemi_sheet=None, lat4_sheet=None):
        """
        Creates charts of mean aod data and appends it to the appropriate datasheet
        :param gm_sheet: Global mean data openpyxl sheet
        :param hemi_sheet: Hemispheric data openpyxl sheet
        :param lat4_sheet: Four latitude regional data openpyxl sheet
        :return: None
        """
        chart_width = 30
        chart_height = 15
        line_width = 5

        if gm_sheet is not None:
            # Create the chart
            chart = ScatterChart(scatterStyle='line')
            # Setup the data series
            time = Reference(gm_sheet, min_col=1, min_row=2, max_col=1, max_row=1+len(self.time))
            data = Reference(gm_sheet, min_col=2, min_row=1, max_col=2, max_row=1+len(self.time))
            line = Series(data, xvalues=time, title_from_data=True)
            # Change the thickness of the line
            line.graphicalProperties.line = LineProperties(w=line_width)
            # Add the line to the chart
            chart.append(line)
            # Add chart titles
            chart.title = "Global Mean SAOD at 550nm"
            chart.x_axis.title = ("Time (" + self.data.time_units + ")")
            chart.y_axis.title = "Mean SAOD at lambda = 550nm"
            # Format chart axes
            chart.x_axis.scaling.min = np.min(self.time)
            chart.x_axis.scaling.max = np.max(self.time)
            # Adjust chart size
            chart.width = chart_width
            chart.height = chart_height
            # Add chart to the sheet
            gm_sheet.add_chart(chart, "D1")

        if hemi_sheet is not None:
            # Create the chart
            chart = ScatterChart(scatterStyle='line')
            # Setup the first data series
            time = Reference(hemi_sheet, min_col=1, min_row=2, max_col=1, max_row=1 + len(self.time))
            data = Reference(hemi_sheet, min_col=2, min_row=1, max_col=2, max_row=1+len(self.time))
            line = Series(data, xvalues=time, title_from_data=True)
            # Adjust the line thickness
            line.graphicalProperties.line = LineProperties(w=line_width)
            # Add the line to the chart
            chart.append(line)
            # Repeat for the second data series
            data = Reference(hemi_sheet, min_col=3, min_row=1, max_col=3, max_row=1 + len(self.time))
            line = Series(data, xvalues=time, title_from_data=True)
            line.graphicalProperties.line = LineProperties(w=line_width)
            chart.append(line)
            # Add chart titles
            chart.title = "Hemispheric Mean SAOD at 550nm"
            chart.x_axis.label = ("Time (" + self.data.time_units + ")")
            chart.y_axis.label = "Mean SAOD at lambda = 550nm"
            # Format the axes
            chart.x_axis.scaling.min = np.min(self.time)
            chart.x_axis.scaling.max = np.max(self.time)
            # Adjust the chart size
            chart.width = chart_width
            chart.height = chart_height
            # Add chart to the sheet
            hemi_sheet.add_chart(chart, "E1")

        if lat4_sheet is not None:
            # Create the chart
            chart = ScatterChart(scatterStyle='line')
            # Extract the time data
            time = Reference(lat4_sheet, min_col=1, min_row=2, max_col=1, max_row=1 + len(self.time))
            # Add data to chart from columns 2-5
            for i in range(2, 6):
                # Extract the data
                data = Reference(lat4_sheet, min_col=i, min_row=1, max_col=i, max_row=1 + len(self.time))
                line = Series(data, xvalues=time, title_from_data=True)
                # Adjust line thickness
                line.graphicalProperties.line = LineProperties(w=line_width)
                # Add the line to the chart
                chart.append(line)
            # Setup chart titles
            chart.title = "Four LatBand Mean SAOD at 550nm"
            chart.x_axis.label = ("Time (" + self.data.time_units + ")")
            chart.y_axis.label = "Mean SAOD at lambda = 550nm"
            # Format the axes
            chart.x_axis.scaling.min = np.min(self.time)
            chart.x_axis.scaling.max = np.max(self.time)
            # Adjust the chart size
            chart.width = chart_width
            chart.height = chart_height
            # Add the chart to the sheet
            lat4_sheet.add_chart(chart, "G1")


def convert_nc_to_excel(filename, aod550=False):
    nc_data = ReadEvaNC(file_in=filename, aod550=aod550)
    writer = WriteData(nc_data)
    writer.toExcel()
    return


def write_to_nc(filename, time, latitude, aod550, metadata, time_unit=None, lat_unit=None, time_desc=None):
    WriteEvaNC(filename, time, latitude, aod550, metadata, time_unit=time_unit, lat_unit=lat_unit, time_desc=time_desc)
    return


