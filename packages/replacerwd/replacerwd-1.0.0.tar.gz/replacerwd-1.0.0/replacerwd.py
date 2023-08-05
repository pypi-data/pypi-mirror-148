#conding=utf8 pyt
from tkinter import*
from tkinter import messagebox
import tkinter;
from tkinter import filedialog
import os
import openpyxl;
from openpyxl.styles import PatternFill
from tkinter import ttk
import docx
from docx.shared import RGBColor
import win32com
from win32com.client import Dispatch

num = 0
gap = 33
begin_y = 140
is_replacing = False
target_path=""
suffix_array=[".xlsx",".docx"] 
target_file_type = suffix_array[0]

lab_index_array = []
edit_old_array = []
lab_arrows_array = []
edit_news_array = []
btn_del_array = []
list_color_box_array = []
fg_color_names = ["无","红色","绿色","蓝色","黄色","巧克力色","紫色"]
fg_color_value = ["","FF0000","00FF00","0000FF","FFFF00","5C3317","9F5F9F"]

def hex_to_rgb(value):
    # value = value.lstrip("#")
    lv = len(value)
    return tuple(int(value[i:i + lv // 3],16) for i in range(0,lv, lv//3))


def replaceWord(word_full_path,word_name):
    
    # 如果使用word
    exec_tool = 'Word.Application'
    # 指示运行的版本，如果是WPS应修改为
    word = win32com.client.Dispatch(exec_tool)
    # 在后台运行程序
    word.Visible = 0  # 后台运行，不显示
    # 运行过程不警告
    word.DisplayAlerts = 0  # 不警告
    # 打开word文档
    doc = word.Documents.Open(word_full_path)
    maxNum = len(lab_index_array)
    for c in range(maxNum):
        oldStr = edit_old_array[c].get()
        newStr = edit_news_array[c].get()
        word.Selection.Find.ClearFormatting()
        word.Selection.Find.Replacement.ClearFormatting()
        word.Selection.Find.Execute(oldStr, False, False, False, False, False, True, 1, True, newStr, 2)
        pass
    doc.SaveAs(word_full_path)
    print("查找替换word文档："+word_name+" 完毕,如果需要替换的话已经替换完成")
    pass

#替换单个word文件
def replaceWordContent(word_full_path,word_name):
    # print("修改word文档："+word_name)
    maxNum = len(lab_index_array)
    doc = docx.Document(word_full_path)
    # color_index = 1
    # color_rgb = hex_to_rgb(fg_color_value[color_index])
    print("maxNum=",maxNum)
    need_change = False
    all_paragraphs = doc.paragraphs
    for pa in all_paragraphs:
        for run in pa.runs:
            run_texts = run.text
            print("run_texts="+run_texts)
            for c in range(maxNum):
                oldStr = edit_old_array[c].get()
                newStr = edit_news_array[c].get()
                print("oldStr="+oldStr+",newStr="+newStr)
                if(oldStr in run_texts):
                    print("需要修改字符串")
                    need_change = True
                    run_texts = run_texts.replace(oldStr,newStr,50)
                    run.text = run_texts
                    color_index = list_color_box_array[c].current()
                    if(color_index>0):
                    #   color_rgb = hex_to_rgb(fg_color_value[color_index])
                    #   run.font.color.rgb = RGBColor(color_rgb[0], color_rgb[1], color_rgb[0])
                      pass
                    pass
                pass
            pass
        pass
    if(need_change):
        print("修改word文档："+word_name)
    doc.save(word_full_path)
    pass

#替换单个execl文件
def replaceExcelContent(excel_full_path,excel_name):
    book = openpyxl.load_workbook(excel_full_path)
    sheetNames = book.sheetnames
    maxNum = len(lab_index_array)
    # print('sheetNames:',sheetNames)
    need_change = False
    has_replace_str_array = []
    for sheet in sheetNames:
            sheetI = book[sheet]
            maxHang = sheetI.max_row
            maxLie = sheetI.max_column
            for a in range(maxHang):
                for b in range(maxLie):
                    item = sheetI.cell(row=a+1, column=b+1).value
                    if(isinstance(item, str)):
                        for c in range(maxNum):
                            oldStr = edit_old_array[c].get()
                            newStr = edit_news_array[c].get()
                            if(oldStr in item):
                                need_change = True
                                color_index = list_color_box_array[c].current()
                                # color_name = list_color_box_array[c].get()
                                # print("color_index="+str(color_index))
                                item = item.replace(oldStr,newStr,20)
                                if(not oldStr in has_replace_str_array):
                                    has_replace_str_array.append(oldStr)
                                    pass
                                # show_tip = "正在将表 "+excel_name+" 里所有 "+oldStr+" 替换成 "+ newStr
                                # print(show_tip)
                                if(color_index>0):
                                    color_fill = PatternFill(fill_type='solid',fgColor=fg_color_value[color_index])
                                    sheetI.cell(a+1, b+1,item).fill = color_fill
                                else:
                                    sheetI.cell(a+1, b+1,item)
                                pass
                            pass
                        pass
                        # if(oldStr in item):
                        #     item = item.replace(oldStr,newStr,100)
                        #     sheetI.cell(a+1, b+1,item).fill = color_fill
    book.save(excel_full_path)
    if(need_change):
        print(excel_name+" 替换了： ",has_replace_str_array)

#批量替换文件夹所有文件
def replace_target_path(dirname):
    # print("dirname:",dirname) #当前主目录
    # print("正在将所有 "+oldStr+" 替换成 "+ newStr)
    # print("target_file_type="+target_file_type)
    for maindir, subdir, file_name_list in os.walk(dirname):

        # print("1:",maindir) #当前主目录
        # print("2:",subdir) #当前主目录下的所有目录
        # print("3:",file_name_list)  #当前主目录下的所有文件

        for filename in file_name_list:
            apath = os.path.join(maindir, filename)#合并成一个完整路径
            # print("filename="+filename)
            ext = os.path.splitext(apath)[1]  # 获取文件后缀 [0]获取的是除了文件名以外的内容
            # print("ext="+ext)
            if ext in suffix_array and ext==target_file_type:
                # print("文件路径 = "+apath)
                # print("正在将表 "+filename+" 里所有 "+oldStr+" 替换成 "+ newStr)
                if(target_file_type==".xlsx"):
                    replaceExcelContent(apath,filename)
                elif(target_file_type==".docx"):
                    # replaceWordContent(apath,filename)
                    replaceWord(apath,filename)
                    pass
                pass
                
                # replaceExcelContent(apath,filename)
    # print("所有 "+oldStr+" 替换成 "+ newStr +" 完毕！！！")
    pass

def start_replace():
    global target_path
    global is_replacing
    global target_file_type
    target_path = edit_search.get()
    target_file_type = combox_file_type.get()
    # print("要替换的文件类型是："+target_file_type)
    if(not os.path.exists(target_path)):
        messagebox.showinfo("出错了","路径不存在，请选择正确路径")
        return
    # print("要替换的文件夹是："+target_path)
    curLength = len(lab_index_array)
    if(curLength==0):
        messagebox.showinfo("出错了","请新增要替换的条目")
        return
    print("当前要替换的条目数量",curLength)
    is_replacing = True
    messagebox.showinfo("准备开始","确认要开始替换吗？点击确认之后请耐心等待替换完成，不要重复操作，好了会弹框告诉你哦！")
    replace_target_path(target_path)
    print("所有替换执行完毕")
    messagebox.showinfo("搞定",target_path+"里面的所有.xlsx已经替换完毕")
    pass

#删除一列
def del_item(itemIndex):
    global num
    # print("删除的下标是：",itemIndex)
    # print(edit_old_array)
    lab_index_array[itemIndex].destroy()
    del lab_index_array[itemIndex]
    # print(lab_index_array)
    edit_old_array[itemIndex].destroy()
    del edit_old_array[itemIndex]
    lab_arrows_array[itemIndex].destroy()
    del lab_arrows_array[itemIndex]
    edit_news_array[itemIndex].destroy()
    del edit_news_array[itemIndex]
    btn_del_array[itemIndex].destroy()
    del btn_del_array[itemIndex]
    list_color_box_array[itemIndex].destroy()
    del list_color_box_array[itemIndex]

    remainLength = len(lab_index_array)
    for i in range(remainLength):
        lab_index_array[i].place(x=100, y=(begin_y+i*gap),height=20)
        lab_index_array[i]["text"] = i
        edit_old_array[i].place(x=160, y=(begin_y+i*gap),height=20)
        lab_arrows_array[i].place(x=380, y=(begin_y+i*gap),height=20)
        edit_news_array[i].place(x=460, y=(begin_y+i*gap),height=20)
        btn_del_array[i].place(x=850, y=(begin_y+i*gap),height=20)
        list_color_box_array[i].place(x=700, y=(begin_y+i*gap))
        pass
    num = num - 1
    pass

#打开文件夹
def open_file():
    path = filedialog.askdirectory()
    edit_search.delete(0, END)  # 删除所有值
    edit_search.insert(0,path)
    # print("path=",path)
    pass

#新增一条替换记录
def add_item():
    global num
    global gap
    global lab_index_array
    global edit_old_array
    global lab_arrows_array
    global edit_news_array
    global btn_del_array
    global list_color_box_array
    # print("num=",num)
    lab_index = Label(root,text=num,font=('宋体',20))
    lab_index.place(x=100, y=(begin_y+num*gap),height=20)
    lab_index_array.append(lab_index)
    edit_old = tkinter.Entry(root)
    edit_old.place(x=160, y=(begin_y+num*gap),height=20)
    edit_old_array.append(edit_old)
    lab_arrow = Label(root,text="->",font=('宋体',20))
    lab_arrow.place(x=380, y=(begin_y+num*gap),height=20)
    lab_arrows_array.append(lab_arrow)
    edit_new = tkinter.Entry(root)
    edit_new.place(x=460, y=(begin_y+num*gap),height=20)
    edit_news_array.append(edit_new)
    #删除按钮
    btnDel = Button(root,text="删除",command=lambda:del_item(lab_index.cget("text")))
    btnDel.place(x=850, y=(begin_y+num*gap),height=20)
    btn_del_array.append(btnDel)
    #颜色滚动列表
    combobox = ttk.Combobox(
      master=root, # 父容器
      height=10, # 高度,下拉显示的条目数量
      width=10, # 宽度
      state="readonly", # 设置状态 normal(可选可输入)、readonly(只可选)、 disabled
      cursor="arrow", # 鼠标移动时样式 arrow, circle, cross, plus...
      font=("", 12), # 字体
      values=fg_color_names, # 设置下拉框的选项
      )
    combobox.current(0)
    combobox.place(x=700, y=(begin_y+num*gap))
    list_color_box_array.append(combobox)
    num = num +1
    pass

def open_entrance():
    # 进入消息循环
    global root
    root = Tk()
    root.geometry('960x640+600+200')
    root.title("excel批量替换工具")
    #查找位置
    lab_search = Label(root,text="查找位置:",font=('宋体',16))
    lab_search.place(x=10, y=20)
    #输入框底图
    # img_edit_Bg_open = Image.open("./imgs/text_bg.png")
    # img_edit_bg_png = ImageTk.PhotoImage(img_edit_Bg_open)
    # lab_img_edit:tkinter.Label = tkinter.Label(root,image = img_edit_bg_png)
    # lab_img_edit.place(x=200,y=20 ,width=400,height=35)
    #输入框
    global edit_search
    edit_search = tkinter.Entry(root)
    edit_search.insert(0,"点击左边打开按钮选择文件夹，或者手动输入文件夹路径")
    edit_search.place(x=210,y=22.5,width=380,height=30)
    #打开按钮
    btnFile = Button(root,text="打开",command=open_file)
    btnFile.place(x=150,y=22.5)
    #文本替换
    lab_search = Label(root,text="文本替换:",font=('宋体',20))
    lab_search.place(x=10, y=60)
    #文件类型
    lab_file_type = Label(root,text="文件类型:",font=('宋体',14))
    lab_file_type.place(x=150, y=65)
    #文件类型下拉框
    global combox_file_type
    combox_file_type = ttk.Combobox(
      master=root, # 父容器
      height=10, # 高度,下拉显示的条目数量
      width=10, # 宽度
      state="readonly", # 设置状态 normal(可选可输入)、readonly(只可选)、 disabled
      cursor="arrow", # 鼠标移动时样式 arrow, circle, cross, plus...
      font=("", 12), # 字体
      values=suffix_array, # 设置下拉框的选项
      )
    combox_file_type.current(0)
    combox_file_type.place(x=250, y=65)


    #新增按钮
    btnFile = Button(root,text="新增",command=add_item)
    btnFile.place(x=400,y=60)
    #开始批量替换按钮
    btnReplace = Button(root,text="开始批量替换",command=start_replace)
    btnReplace.place(x=550,y=60)
    #原文本
    lab_old = Label(root,text="原文本:",font=('宋体',16))
    lab_old.place(x=160, y=100)
    #新文本
    lab_new = Label(root,text="新文本:",font=('宋体',16))
    lab_new.place(x=460, y=100)


    #颜色标注
    lab_color = Label(root,text="颜色标注:",font=('宋体',16))
    lab_color.place(x=700, y=100)

    root.mainloop()
    pass
